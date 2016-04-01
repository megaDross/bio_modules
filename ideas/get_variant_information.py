import openpyxl

# variant found spits out the last one
variant_header = {}
mutation_header = {}

def open_variant_validation_spreadsheets(excel_sheet="All_Yale_&_UK_Variants.xlsx",
                                         tab1='All_Variants', tab2="Mutations ID",
                                         header_row=2):
    '''Open the Yale & UK TAAD variant validation workbook and
       the relevant variant and mutation sheets
    '''    
    # open spreadsheets and tabs within
    wb = openpyxl.load_workbook(excel_sheet,data_only=True)
    variants = wb.get_sheet_by_name(tab1) # or wb["All_variants"]
    mutations = wb[tab2]
    
    # use the headers of the given tabs to create a dictionary
    create_header_dicts(mutations,mutation_header)
    create_header_dicts(variants,variant_header)
    
    return variants, mutations
          
          
def create_header_dicts(tab,dict_name,header_row=2):
    '''Store each entry in a header in a given spreadsheet as
       a key with no value. 
    '''
    for column_number in range(1,100):
        header = tab.cell(row=2,column=column_number).value
        if header is None:
            break
        dict_name[header] = ''
    
  
                          
def get_row(query,sheet,column_num=1):
    ''' Search the database/sheet and match with the query/variant_alias
        if found, output its row number in the database
    '''
    for row_number in range(1,500):
        if query == sheet.cell(row=row_number,column=column_num).value:
            return row_number 



def get_variant_info(row,sheet,dic):
    ''' Extract information associated with the query inputted in get_row() 
        from the database and append it to the items in variant_header dict
    '''
    for i in range(1,100):
        column = sheet.cell(row=2,column=i).value
        get_info =sheet.cell(row=row,column=i).value
        if column in dic:
            dic[column]=get_info
            if dic.get(column) is None:
                dic[column] = "-"
    return dic
    

