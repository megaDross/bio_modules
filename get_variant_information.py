import openpyxl

variant_header = {
"Sample_Name": '', "Gene": '',
"Exon_No.": '', "HGVSc": '',
"HGVSp": '', "Variant_Position": '',
"Allele_Balance": '', "Allele_Depth(REF)": '',
"Allele_Depth(ALT)": '', "Allele_Frequency_ESP(%)": '',
"Allele_Frequency_ExAC(%)": '', "Allele_Frequency_dbSNP(%)": '',
"Variant_Found": '', "Mutation_ID": '',"Reason for Variant class change": '',
"Category":'', "Variant_Alias":''
}

mutation_header = {
"Report Variant class field": '', "First Published": '',
"Reason for Variant class change": '', "Date of variant class change": '',
"Variant Class": ''
}



def open_variant_validation_spreadsheets():
    '''Open the Yale & UK TAAD variant validation workbook and
       the relevant variant and mutation sheets
    '''    
    
    wb = openpyxl.load_workbook("All_Yale_&_UK_Variants.xlsx",data_only=True)
    variants = wb.get_sheet_by_name('All_Variants') # or wb["All_variants"]
    mutations = wb["Mutations ID"]
    return variants, mutations
          
                
                          
def get_row(query,sheet,column_num=1):
    ''' Search the database/sheet and match with the query/variant_alias
        if found, output its row number in the database
    '''
    for row_number in range(1,500):
        if query == sheet.cell(row=row_number,column=column_num).value:
            return row_number 



def get_variant_info(row,sheet,dic):
    ''' Extract information associated with the query inputted in get_row() 
        from the database and append it to the items in variant_header dict
    '''
    for i in range(1,100):
        column = sheet.cell(row=2,column=i).value
        get_info =sheet.cell(row=row,column=i).value
        if column in dic:
            dic[column]=get_info
            if dic.get(column) is None:
                dic[column] = "-"
    return dic